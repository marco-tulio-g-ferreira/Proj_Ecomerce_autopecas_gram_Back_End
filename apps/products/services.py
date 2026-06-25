import re
import csv
import io
import requests
import os
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.db import transaction
from django.core.cache import cache

# Imports locais conforme sua estrutura
from .models import Part, Category, LocalEstoque, Estoque, RevisaoDados
from .constants import CATEGORIAS_MAPEAMENTO, MAX_PRICE_THRESHOLD

# --- CONFIGURAÇÃO DA API SICOOB ---
SICOOB_AUTH_URL = "https://api.sandbox.sicoob.com.br/oauth2/token"
SICOOB_API_URL = "https://api.sandbox.sicoob.com.br/pix/api/v2/cob"
CLIENT_ID = os.getenv('SICOOB_CLIENT_ID')
CLIENT_SECRET = os.getenv('SICOOB_CLIENT_SECRET')
CERT_PATH = os.getenv('SICOOB_CERT_PATH')

# --- Lógica de Autenticação Dinâmica (Token) ---
def get_sicoob_token():
    """
    Busca o token no cache ou solicita um novo ao Sicoob caso tenha expirado.
    """
    token = cache.get("sicoob_access_token")
    if token:
        return token

    try:
        response = requests.post(
            SICOOB_AUTH_URL,
            data={'grant_type': 'client_credentials', 'scope': 'cob.write cob.read'},
            auth=(CLIENT_ID, CLIENT_SECRET),
            cert=CERT_PATH 
        )
        response.raise_for_status()
        data = response.json()
        
        token = data.get('access_token')
        expires_in = int(data.get('expires_in', 3600))
        
        # Salva no cache com margem de segurança de 60 segundos
        cache.set("sicoob_access_token", token, timeout=expires_in - 60)
        return token
    except Exception as e:
        print(f"Erro ao obter token Sicoob: {e}")
        return None

def gerar_cobranca_pix(valor):
    """
    Gera cobrança PIX usando o token dinâmico configurado.
    """
    token = get_sicoob_token()
    if not token:
        raise Exception("Falha na autenticação com o Sicoob.")

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "valor": {"original": str(valor)},
        "chave": os.getenv('PIX_CHAVE_KEY'),
        "solicitacaoPagador": "Pagamento de produto"
    }

    try:
        response = requests.post(SICOOB_API_URL, json=payload, headers=headers, cert=CERT_PATH)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        raise Exception(f"Erro ao comunicar com API Sicoob: {str(e)}")

# --- Lógica de Importação ---
def atualizar_progresso(task_id, prog, msg, success=None, err=None):
    dados = {"progress": prog, "message": msg}
    if success is not None: dados["sucessos"] = success
    if err is not None: dados["erros"] = err
    cache.set(f"import_progress_{task_id}", dados, timeout=3600)

def limpar_valor(valor_bruto):
    """Sanitiza preços vindo do SIC-Net."""
    if valor_bruto is None or valor_bruto == "":
        return Decimal("0.00")
    
    valor_str = str(valor_bruto).replace("R$", "").strip()
    valor_str = re.sub(r'[^\d,.]', '', valor_str)
    
    if ',' in valor_str and '.' in valor_str:
        if valor_str.find('.') < valor_str.find(','):
            valor_str = valor_str.replace('.', '').replace(',', '.')
        else:
            valor_str = valor_str.replace(',', '')
    elif ',' in valor_str:
        valor_str = valor_str.replace(',', '.')
        
    partes = valor_str.split('.')
    if len(partes) > 2:
        valor_str = "".join(partes[:-1]) + "." + partes[-1]
    
    try:
        return Decimal(valor_str)
    except (InvalidOperation, ValueError, Exception):
        return Decimal("0.00")

def validar_preco_realista(preco):
    """Retorna True se o preço for válido e realista."""
    return 0 < preco <= MAX_PRICE_THRESHOLD

def limpar_quantidade(valor_bruto):
    """Trata strings como '0,4' ou '4,8' convertendo para inteiro."""
    if valor_bruto is None or valor_bruto == "":
        return 0
    valor_str = str(valor_bruto).replace(",", ".").strip()
    try:
        return int(float(valor_str))
    except (ValueError, TypeError):
        return 0

def obter_categoria_pelo_nome(nome_produto):
    """Classificação inteligente usando mapeamento centralizado."""
    nome = nome_produto.upper()
    for categoria, termos in CATEGORIAS_MAPEAMENTO.items():
        if any(term in nome for term in termos):
            cat_obj, _ = Category.objects.get_or_create(name=categoria)
            return cat_obj
    cat_obj, _ = Category.objects.get_or_create(name="Geral")
    return cat_obj

def processar_importacao_background(file_content, filename, task_id):
    local_padrao, _ = LocalEstoque.objects.get_or_create(
        nome_local="Depósito Central", 
        defaults={'prateleira': 'A1', 'box': '01'}
    )

    sucessos = 0
    erros = 0
    linhas = []

    try:
        if filename.lower().endswith('.csv'):
            try: reader = csv.reader(io.StringIO(file_content.decode('utf-8')))
            except UnicodeDecodeError: reader = csv.reader(io.StringIO(file_content.decode('latin-1')))
            next(reader, None)
            linhas = list(reader)
        else:
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
            aba = wb.active
            linhas = [list(linha) for linha in aba.iter_rows(min_row=2, values_only=True)]
    except Exception as e:
        atualizar_progresso(task_id, 100, f"Erro ao ler arquivo: {str(e)}", 0, 1)
        return

    total_linhas = len(linhas)
    
    for index, linha in enumerate(linhas, start=1):
        try:
            if not linha or len(linha) < 2 or (linha[1] is None): continue
            
            sku_da_planilha = str(linha[0]).strip() if linha[0] is not None else ""
            nome_produto = str(linha[1]).strip()
            
            # Limpeza
            qtd_final = limpar_quantidade(linha[2])
            preco_final = limpar_valor(linha[3])
            
            # --- VALIDAÇÃO COM REGISTRO DE ERRO (REVISÃO) ---
            if not validar_preco_realista(preco_final):
                RevisaoDados.objects.create(
                    nome_produto=nome_produto,
                    sku_tentativa=sku_da_planilha,
                    preco_tentativa=preco_final,
                    motivo=f"Preço fora do limite: R$ {preco_final}"
                )
                erros += 1
                if index % 5 == 0 or index == total_linhas:
                    percentual = int((index / total_linhas) * 100)
                    atualizar_progresso(task_id, percentual, f"Processando... ({erros} itens para revisão)", sucessos, erros)
                continue 
            
            categoria_obj = obter_categoria_pelo_nome(nome_produto)

            with transaction.atomic():
                if sku_da_planilha and sku_da_planilha.lower() != "none":
                    part, _ = Part.objects.update_or_create(
                        sku=sku_da_planilha, 
                        defaults={
                            'category': categoria_obj, 
                            'name': nome_produto, 
                            'codigo_oem': f"OEM-{sku_da_planilha}"
                        }
                    )
                else:
                    part, _ = Part.objects.update_or_create(
                        name=nome_produto,
                        defaults={
                            'category': categoria_obj,
                            'sku': f"GEN-{hash(nome_produto) & 0xffffffff}",
                            'codigo_oem': "SEM-OEM"
                        }
                    )
                
                Estoque.objects.update_or_create(
                    produto=part, 
                    defaults={'local': local_padrao, 'quantidade': qtd_final, 'preco': preco_final}
                )
            sucessos += 1
        except Exception as e:
            print(f"Erro na linha {index}: {e}")
            erros += 1
        
        if index % 5 == 0 or index == total_linhas:
            percentual = int((index / total_linhas) * 100)
            atualizar_progresso(task_id, percentual, f"Processando linha {index} de {total_linhas}...", sucessos, erros)

    atualizar_progresso(task_id, 100, "Concluído!", sucessos, erros)