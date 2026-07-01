from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
import os
from decimal import Decimal
from apps.products.models import Part, Category, LocalEstoque, Estoque

class Command(BaseCommand):
    help = 'Importa dados de peças a partir de um arquivo Excel de forma otimizada'

    def add_arguments(self, parser):
        parser.add_argument('caminho_excel', type=str, help='Caminho do arquivo na raiz')
        parser.add_argument('--categoria', type=str, default='Geral', help='Nome da categoria')

    def handle(self, *args, **kwargs):
        caminho_excel = kwargs['caminho_excel']
        nome_cat = kwargs['categoria']

        if not os.path.exists(caminho_excel):
            self.stdout.write(self.style.ERROR(f"❌ Arquivo '{caminho_excel}' não encontrado."))
            return

        # 1. Preparação: Cache em memória
        self.stdout.write("📥 Carregando dados existentes para o cache...")
        categoria_obj, _ = Category.objects.get_or_create(name=nome_cat)
        local_padrao, _ = LocalEstoque.objects.get_or_create(
            nome_local="Depósito Central",
            defaults={'prateleira': 'A1', 'box': '01'}
        )

        parts_map = {p.sku: p for p in Part.objects.all()}
        stocks_map = {e.produto_id: e for e in Estoque.objects.filter(local=local_padrao)}

        # 2. Leitura rápida do Excel
        wb = load_workbook(caminho_excel, data_only=True, read_only=True)
        aba = wb.active

        to_create_parts = []
        to_update_parts = []
        to_create_stocks = []
        to_update_stocks = []
        
        # Mapeia SKUs que precisam ter estoque criado logo após o bulk_create de parts
        sku_to_stock_data = {}

        self.stdout.write("⚙️ Processando planilha...")
        
        for index, linha in enumerate(aba.iter_rows(min_row=2, values_only=True), start=2):
            if not linha or linha[1] is None:
                continue

            sku = str(linha[0]).strip() if linha[0] is not None else None
            nome_produto = str(linha[1]).strip()
            qtd = int(float(linha[2] or 0))
            valor_bruto = str(linha[3] or "0").replace("R$", "").replace(".", "").replace(",", ".").strip()
            preco = Decimal(valor_bruto) if valor_bruto else Decimal("0.00")

            # Lógica de Part
            if sku and sku in parts_map:
                part = parts_map[sku]
                if part.name != nome_produto:
                    part.name = nome_produto
                    to_update_parts.append(part)
                
                # Se a peça já existe, verificamos o estoque
                if part.id in stocks_map:
                    stock = stocks_map[part.id]
                    stock.quantidade = qtd
                    stock.preco = preco
                    to_update_stocks.append(stock)
                else:
                    to_create_stocks.append(Estoque(produto=part, local=local_padrao, quantidade=qtd, preco=preco))
            else:
                # Part nova
                part = Part(sku=sku or f"GEN-{index}", name=nome_produto, category=categoria_obj)
                to_create_parts.append(part)
                # Guardamos os dados do estoque para criar depois que o Part tiver ID
                sku_to_stock_data[part.sku] = {'qtd': qtd, 'preco': preco}
                parts_map[part.sku] = part 

        # 3. Execução em Bloco (Bulk)
        self.stdout.write("💾 Salvando no banco de dados...")
        with transaction.atomic():
            # A. Salva/Atualiza Parts
            if to_create_parts:
                Part.objects.bulk_create(to_create_parts)
                # Recarrega os objetos criados para obter os IDs gerados pelo banco
                new_parts = Part.objects.filter(sku__in=[p.sku for p in to_create_parts])
                for p in new_parts:
                    stock_data = sku_to_stock_data.get(p.sku)
                    if stock_data:
                        to_create_stocks.append(Estoque(
                            produto=p, local=local_padrao, 
                            quantidade=stock_data['qtd'], preco=stock_data['preco']
                        ))
            
            if to_update_parts:
                Part.objects.bulk_update(to_update_parts, ['name'])

            # B. Salva/Atualiza Estoques
            if to_create_stocks:
                Estoque.objects.bulk_create(to_create_stocks)
            if to_update_stocks:
                Estoque.objects.bulk_update(to_update_stocks, ['quantidade', 'preco'])

        self.stdout.write(self.style.SUCCESS("✅ Importação concluída com sucesso!"))